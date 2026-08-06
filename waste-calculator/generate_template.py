"""廃棄物コネクト 案件量計算ツール - Excelテンプレート生成スクリプト。

再生成する場合: python3 generate_template.py
出力: waste_calculator_template.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3D24")
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=10)
INPUT_FONT = Font(name=FONT_NAME, color="0000FF", size=10)
FORMULA_FONT = Font(name=FONT_NAME, color="000000", size=10)
NOTE_FONT = Font(name=FONT_NAME, italic=True, color="6B6F65", size=9)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F3D24")
SUBTITLE_FONT = Font(name=FONT_NAME, size=10, color="6B6F65")
ASSUMPTION_FILL = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="D9DCD3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TYPE_LIST = ["可燃ごみ", "生ごみ", "不燃ごみ", "瓶", "缶", "ペットボトル", "段ボール", "その他"]

wb = openpyxl.Workbook()

# ---------------------------------------------------------------- 使い方
ws_help = wb.active
ws_help.title = "使い方"
ws_help.sheet_view.showGridLines = False
ws_help.column_dimensions["A"].width = 2
ws_help.column_dimensions["B"].width = 100

r = 2
ws_help.cell(r, 2, "廃棄物コネクト 案件量計算ツール").font = TITLE_FONT
r += 1
ws_help.cell(r, 2, "ヒアリング内容から月間の袋数・容量・概算料金を自動計算するテンプレートです。").font = SUBTITLE_FONT
r += 2
lines = [
    ("① どちらのシートを使うか", True),
    ("「かんたん入力」：種別・数量・頻度だけで月間袋数をサッと出したいとき。", False),
    ("「詳細入力」：単位・袋サイズ・重量・単価まで含めて月間容量や概算料金も試算したいとき。", False),
    ("どちらも青字のセルのみ入力してください（黒字は自動計算の数式です）。", False),
    ("", False),
    ("② 入力のしかた", True),
    ("・案件名：案件ごとに分かる名前を入力（複数案件を1シートに並べてOK）", False),
    ("・種別：可燃ごみ／生ごみ／不燃ごみ／瓶／缶／ペットボトル／段ボール／その他 からドロップダウンで選択", False),
    ("・頻度(週◯回)：収集の頻度。例:「1日1〜2袋、週5回」→数量1〜2・頻度5、「週に1〜2袋」→数量1〜2・頻度1", False),
    ("・数量min／数量max：「1〜2袋」のように幅がある場合は両方入力。「2袋」のように1つだけの場合は数量maxを空欄にすればOK（数量minの値だけで計算されます）。", False),
    ("・（詳細入力のみ）重量(kg/個)：料金計算のもとになる重量。45L≈5kg・70L≈8kgの実測値をもとにした目安を仮入力していますが、必要に応じて修正してください（段ボールは1枚≈0.3kgが目安）。", False),
    ("・（詳細入力のみ）単価(円/kg)：1kgあたりの料金。可燃40円・不燃100円・段ボール35円を仮入力済み（実際の契約単価に合わせて修正してください）", False),
    ("", False),
    ("③ 前提", True),
    ("1ヶ月あたりの週数は各シートのC1セルで設定（既定値 4.35週 ＝ 365日 ÷ 12ヶ月 ÷ 7日）。", False),
    ("数量が明記されていない項目（例:「不燃物 週2日」のみで数量記載なし）は 1回1袋 と仮定しています。実数が分かり次第、数量欄を修正してください。", False),
    ("概算料金＝月間量×重量(kg/個)×単価(円/kg)で計算しています。重量が未入力の項目は料金計算から除外されます。", False),
    ("初回手数料（詳細入力シートC2セル、既定 ¥3,000）は新規契約の案件ごとに一律で加算する費用です。ご提示金額＝その案件の月間概算料金の合計＋初回手数料、として手動で合算してください。", False),
]
for text, bold in lines:
    c = ws_help.cell(r, 2, text)
    c.font = Font(name=FONT_NAME, bold=True, size=11, color="1F3D24") if bold else Font(name=FONT_NAME, size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1


def add_type_validation(ws, col_letter, first_row, last_row):
    dv = DataValidation(type="list", formula1='"{}"'.format(",".join(TYPE_LIST)), allow_blank=True, showDropDown=False)
    dv.showErrorMessage = False
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


# ---------------------------------------------------------------- かんたん入力
ws_s = wb.create_sheet("かんたん入力")
ws_s.sheet_view.showGridLines = False

ws_s.cell(1, 1, "1ヶ月あたりの週数").font = Font(name=FONT_NAME, bold=True, size=10)
w = ws_s.cell(1, 3, 4.35)
w.font = INPUT_FONT
w.fill = ASSUMPTION_FILL
ws_s.cell(1, 4, "※365日÷12ヶ月÷7日で算出した平均週数。必要に応じて変更してください。").font = NOTE_FONT

s_headers = ["案件名", "種別", "数量min(1回あたり)", "数量max(1回あたり)", "頻度(週◯回)", "月間量(平均)"]
s_header_row = 3
for i, h in enumerate(s_headers, start=1):
    c = ws_s.cell(s_header_row, i, h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BORDER
for i, wid in enumerate([26, 12, 16, 16, 12, 14], start=1):
    ws_s.column_dimensions[get_column_letter(i)].width = wid
ws_s.row_dimensions[s_header_row].height = 30

s_samples = [
    # site, type, qmin, qmax, freq
    ["新宿区の新規案件（laidback cafeの例）", "可燃ごみ", 1, 2, 7],
    ["新宿区の新規案件（laidback cafeの例）", "生ごみ", 1, 2, 7],
    ["新宿区の新規案件（laidback cafeの例）", "瓶", 1, 2, 1],
    ["新宿区の新規案件（laidback cafeの例）", "缶", 1, 2, 1],
    ["新宿区の新規案件（laidback cafeの例）", "ペットボトル", 1, 2, 1],
    ["新宿区の新規案件（laidback cafeの例）", "段ボール", 5, 10, 1],
    ["中野区の現行案件の例", "可燃ごみ", 1, 2, 5],
    ["中野区の現行案件の例", "不燃ごみ", 1, 1, 2],
]

s_first_row = s_header_row + 1
s_last_row = s_first_row + 39

for offset, row_data in enumerate(s_samples):
    r = s_first_row + offset
    site, typ, qmin, qmax, freq = row_data
    ws_s.cell(r, 1, site).font = INPUT_FONT
    ws_s.cell(r, 2, typ).font = INPUT_FONT
    ws_s.cell(r, 3, qmin).font = INPUT_FONT
    ws_s.cell(r, 4, qmax).font = INPUT_FONT
    ws_s.cell(r, 5, freq).font = INPUT_FONT

for r in range(s_first_row, s_last_row + 1):
    # F: 月間量(平均) = 数量(1つだけならその値、min〜maxならその平均) * 頻度(週) * 週数
    ws_s.cell(r, 6, f'=IF(C{r}="","",IF(D{r}="",C{r},(C{r}+D{r})/2)*E{r}*$C$1)')
    for col in range(1, 7):
        cell = ws_s.cell(r, col)
        cell.border = BORDER
        if col == 6:
            cell.font = FORMULA_FONT
            cell.number_format = "#,##0.0"

add_type_validation(ws_s, "B", s_first_row, s_last_row)
ws_s.freeze_panes = f"A{s_first_row}"

# ---------------------------------------------------------------- 詳細入力
ws = wb.create_sheet("詳細入力")
ws.sheet_view.showGridLines = False

ws.cell(1, 1, "1ヶ月あたりの週数").font = Font(name=FONT_NAME, bold=True, size=10)
weeks_cell = ws.cell(1, 3, 4.35)
weeks_cell.font = INPUT_FONT
weeks_cell.fill = ASSUMPTION_FILL
ws.cell(1, 4, "※365日÷12ヶ月÷7日で算出した平均週数。必要に応じて変更してください。").font = NOTE_FONT

ws.cell(2, 1, "初回手数料(円)").font = Font(name=FONT_NAME, bold=True, size=10)
setup_fee_cell = ws.cell(2, 3, 3000)
setup_fee_cell.font = INPUT_FONT
setup_fee_cell.fill = ASSUMPTION_FILL
ws.cell(2, 4, "※新規契約時に案件ごと一律で加算。ご提示金額＝月間概算料金の合計＋この金額（手動で合算してください）。").font = NOTE_FONT

headers = ["案件名", "種別", "単位(袋/枚)", "頻度パターン(日/週)", "数量min(1回あたり)",
           "数量max(1回あたり)", "週の収集日数(日パターンのみ)", "袋サイズ(L)", "重量(kg/個)", "単価(円/kg)",
           "月間量(平均)", "月間容量(L)", "月間重量(kg)", "月間概算料金(円)"]
header_row = 3
for i, h in enumerate(headers, start=1):
    c = ws.cell(header_row, i, h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BORDER

widths = [26, 12, 10, 16, 14, 14, 18, 10, 11, 11, 12, 12, 12, 14]
for i, wid in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = wid
ws.row_dimensions[header_row].height = 30

# 重量(kg/個)の目安: 45L≈5kg・70L≈8kgの実測値からの線形換算（袋モノ共通）。段ボールは1枚≈0.3kg。
PRICE_PRESETS = {"可燃ごみ": 40, "不燃ごみ": 100, "段ボール": 35}


def suggest_weight(unit, bagsize, typ):
    if typ == "段ボール":
        return 0.3
    if unit == "袋" and bagsize:
        return round(max(0.0, 0.12 * (bagsize - 45) + 5), 2)
    return ""


sample_rows = [
    # site, type, unit, freq, qmin, qmax, days, bagsize
    ["新宿区の新規案件（laidback cafeの例）", "可燃ごみ", "袋", "日", 1, 2, 7, 45],
    ["新宿区の新規案件（laidback cafeの例）", "生ごみ", "袋", "日", 1, 2, 7, 45],
    ["新宿区の新規案件（laidback cafeの例）", "瓶", "袋", "週", 1, 2, "", 45],
    ["新宿区の新規案件（laidback cafeの例）", "缶", "袋", "週", 1, 2, "", 45],
    ["新宿区の新規案件（laidback cafeの例）", "ペットボトル", "袋", "週", 1, 2, "", 45],
    ["新宿区の新規案件（laidback cafeの例）", "段ボール", "枚", "週", 5, 10, "", ""],
    ["中野区の現行案件の例", "可燃ごみ", "袋", "日", 1, 2, 5, 45],
    ["中野区の現行案件の例", "不燃ごみ", "袋", "日", 1, 1, 2, 45],
]

first_data_row = header_row + 1
last_data_row = first_data_row + 39  # room for extra manual rows below samples

for offset, row_data in enumerate(sample_rows):
    r = first_data_row + offset
    site, typ, unit, freq, qmin, qmax, days, bagsize = row_data
    weight = suggest_weight(unit, bagsize if bagsize else 0, typ)
    price = PRICE_PRESETS.get(typ, "")
    ws.cell(r, 1, site).font = INPUT_FONT
    ws.cell(r, 2, typ).font = INPUT_FONT
    ws.cell(r, 3, unit).font = INPUT_FONT
    ws.cell(r, 4, freq).font = INPUT_FONT
    ws.cell(r, 5, qmin).font = INPUT_FONT
    ws.cell(r, 6, qmax).font = INPUT_FONT
    ws.cell(r, 7, days).font = INPUT_FONT
    ws.cell(r, 8, bagsize).font = INPUT_FONT
    ws.cell(r, 9, weight).font = INPUT_FONT
    ws.cell(r, 10, price).font = INPUT_FONT

for r in range(first_data_row, last_data_row + 1):
    # K: 月間量(平均) = 数量(1つだけならその値、min〜maxならその平均) を頻度パターンに応じて月換算
    qty_expr = f'IF(F{r}="",E{r},(E{r}+F{r})/2)'
    ws.cell(r, 11, f'=IF(E{r}="","",IF(D{r}="日",{qty_expr}*G{r},{qty_expr})*$C$1)')
    # L: 月間容量(L) = 月間量 * 袋サイズ（袋サイズ未入力なら空欄）
    ws.cell(r, 12, f'=IF(OR(K{r}="",H{r}=""),"",K{r}*H{r})')
    # M: 月間重量(kg) = 月間量 * 重量(kg/個)（重量未入力なら空欄）
    ws.cell(r, 13, f'=IF(OR(K{r}="",I{r}=""),"",K{r}*I{r})')
    # N: 月間概算料金 = 月間重量 * 単価（単価未入力なら空欄）
    ws.cell(r, 14, f'=IF(OR(M{r}="",J{r}=""),"",M{r}*J{r})')
    for col in range(1, 15):
        cell = ws.cell(r, col)
        cell.border = BORDER
        if col >= 11:
            cell.font = FORMULA_FONT
            cell.number_format = "#,##0.0"
    ws.cell(r, 14).number_format = "#,##0"

add_type_validation(ws, "B", first_data_row, last_data_row)

dv_unit = DataValidation(type="list", formula1='"袋,枚"', allow_blank=True, showDropDown=False)
dv_unit.showErrorMessage = False
ws.add_data_validation(dv_unit)
dv_unit.add(f"C{first_data_row}:C{last_data_row}")

dv_freq = DataValidation(type="list", formula1='"日,週"', allow_blank=True, showDropDown=False)
dv_freq.showErrorMessage = False
ws.add_data_validation(dv_freq)
dv_freq.add(f"D{first_data_row}:D{last_data_row}")

ws.freeze_panes = f"A{first_data_row}"

wb.save("waste_calculator_template.xlsx")
print("saved waste_calculator_template.xlsx")
